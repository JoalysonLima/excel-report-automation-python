"""Create the final formatted Excel sales report."""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
CLEAN_DATA_FILE = PROJECT_ROOT / "data" / "processed" / "clean_sales_data.xlsx"
SALES_SUMMARY_FILE = PROJECT_ROOT / "data" / "output" / "sales_summary.xlsx"
DATA_QUALITY_FILE = PROJECT_ROOT / "data" / "output" / "data_quality_report.xlsx"
OUTPUT_FILE = PROJECT_ROOT / "data" / "output" / "final_sales_report.xlsx"

TITLE_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
PASS_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
TITLE_FONT = Font(bold=True, color="FFFFFF", size=14)
HEADER_FONT = Font(bold=True, color="000000")
PASS_FONT = Font(color="006100")
FAIL_FONT = Font(color="9C0006")
CURRENCY_FORMAT = '"$"#,##0.00'
PERCENT_FORMAT = "0.0%"
DATE_FORMAT = "yyyy-mm-dd"
MISSING_REGION_LABEL = "Unknown / Missing Region"


def check_input_file(file_path: Path) -> None:
    """Raise a clear error if a required input file is missing."""
    if not file_path.exists():
        raise FileNotFoundError(f"Required input file not found: {file_path}")


def load_report_inputs() -> tuple[pd.DataFrame, dict[str, pd.DataFrame], pd.DataFrame, pd.DataFrame]:
    """Load the cleaned data, sales summaries, and data quality report."""
    for file_path in [CLEAN_DATA_FILE, SALES_SUMMARY_FILE, DATA_QUALITY_FILE]:
        check_input_file(file_path)

    clean_data = pd.read_excel(CLEAN_DATA_FILE)
    sales_summary = pd.read_excel(SALES_SUMMARY_FILE, sheet_name=None)
    quality_summary = pd.read_excel(DATA_QUALITY_FILE, sheet_name="Validation Summary")
    issue_details = pd.read_excel(DATA_QUALITY_FILE, sheet_name="Issue Details")

    return clean_data, sales_summary, quality_summary, issue_details


def build_executive_summary(
    clean_data: pd.DataFrame,
    quality_summary: pd.DataFrame,
    issue_details: pd.DataFrame,
) -> pd.DataFrame:
    """Build a small executive summary table for the first report sheet."""
    total_revenue = clean_data["revenue"].sum()
    total_orders = clean_data["order_id"].nunique()
    total_quantity_sold = clean_data["quantity"].sum()
    average_order_value = total_revenue / total_orders if total_orders else 0

    failed_checks = quality_summary[quality_summary["status"] == "FAIL"]
    passed_checks = quality_summary[quality_summary["status"] == "PASS"]
    quality_pass_rate = (
        len(passed_checks) / len(quality_summary) if len(quality_summary) else 0
    )
    unique_problem_rows = (
        issue_details["row_number"].nunique() if "row_number" in issue_details else 0
    )

    order_dates = pd.to_datetime(clean_data["order_date"], errors="coerce")
    date_range = (
        f"{order_dates.min().date()} to {order_dates.max().date()}"
        if order_dates.notna().any()
        else "No valid dates"
    )

    return pd.DataFrame(
        [
            {"Metric": "Total Revenue", "Value": total_revenue},
            {"Metric": "Total Orders", "Value": total_orders},
            {"Metric": "Total Quantity Sold", "Value": total_quantity_sold},
            {"Metric": "Average Order Value", "Value": average_order_value},
            {"Metric": "Order Date Range", "Value": date_range},
            {"Metric": "Validation Checks Run", "Value": len(quality_summary)},
            {"Metric": "Failed Validation Checks", "Value": len(failed_checks)},
            {"Metric": "Total Issue Occurrences", "Value": len(issue_details)},
            {"Metric": "Unique Problem Rows", "Value": unique_problem_rows},
            {"Metric": "Quality Pass Rate", "Value": quality_pass_rate},
        ]
    )


def build_region_sales_summary(clean_data: pd.DataFrame) -> pd.DataFrame:
    """Build region sales, keeping missing regions visible in the final report."""
    region_data = clean_data.copy()
    region_data["region"] = region_data["region"].fillna("").astype(str).str.strip()
    region_data.loc[region_data["region"] == "", "region"] = MISSING_REGION_LABEL

    return (
        region_data.groupby("region", as_index=False)["revenue"]
        .sum()
        .sort_values("revenue", ascending=False)
    )


def remove_duplicate_missing_checks(quality_summary: pd.DataFrame) -> pd.DataFrame:
    """Remove duplicate missing-field checks from the final report summary."""
    summary = quality_summary.copy()
    check_names = set(summary["check_name"].astype(str))
    duplicate_check_names = []

    for check_name in check_names:
        if not check_name.startswith("Missing values: "):
            continue

        column_name = check_name.replace("Missing values: ", "", 1)
        short_check_name = f"Missing {column_name}"

        if short_check_name in check_names:
            duplicate_check_names.append(short_check_name)

    return summary[~summary["check_name"].isin(duplicate_check_names)].reset_index(
        drop=True
    )


def friendly_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Return a copy of a dataframe with readable column names."""
    renamed = dataframe.copy()
    renamed.columns = [
        str(column_name).replace("_", " ").title() for column_name in renamed.columns
    ]
    return renamed


def write_title(worksheet, title: str, last_column: int) -> None:
    """Add a formatted title row to a worksheet."""
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    title_cell = worksheet.cell(row=1, column=1)
    title_cell.value = title
    title_cell.fill = TITLE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")


def apply_table_formatting(worksheet, header_row: int = 3) -> None:
    """Apply simple professional formatting to a worksheet table."""
    for cell in worksheet[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    worksheet.freeze_panes = f"A{header_row + 1}"
    filter_end_row = worksheet.max_row

    if worksheet.title == "Data Quality":
        for row_index in range(header_row + 1, worksheet.max_row + 1):
            if worksheet.cell(row=row_index, column=1).value == "Issue Details":
                filter_end_row = row_index - 3
                break

    worksheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(worksheet.max_column)}{filter_end_row}"
    )

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 35)


def apply_status_formatting(worksheet, header_row: int = 3) -> None:
    """Format PASS and FAIL status values in the Data Quality sheet."""
    status_column = None

    for cell in worksheet[header_row]:
        if str(cell.value).lower() == "status":
            status_column = cell.column
            break

    if status_column is None:
        return

    for row_index in range(header_row + 1, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_index, column=status_column)
        status = str(cell.value).upper()

        if status == "PASS":
            cell.fill = PASS_FILL
            cell.font = PASS_FONT
        elif status == "FAIL":
            cell.fill = FAIL_FILL
            cell.font = FAIL_FONT


def apply_number_formats(worksheet, header_row: int = 3) -> None:
    """Apply currency, percentage, and date formats based on column names."""
    headers = {
        cell.column: str(cell.value).lower()
        for cell in worksheet[header_row]
        if cell.value is not None
    }

    for column_index, header in headers.items():
        for row_index in range(header_row + 1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)

            if worksheet.title == "Executive Summary" and header == "value":
                metric_cell = worksheet.cell(row=row_index, column=1)
                metric_name = str(metric_cell.value).lower()

                if "rate" in metric_name:
                    cell.number_format = PERCENT_FORMAT
                elif "revenue" in metric_name or "average order value" in metric_name:
                    cell.number_format = CURRENCY_FORMAT
                elif "date" not in metric_name:
                    cell.number_format = "#,##0"
            elif "revenue" in header or "price" in header:
                cell.number_format = CURRENCY_FORMAT
            elif "discount" in header or "rate" in header:
                cell.number_format = PERCENT_FORMAT
            elif "date" in header:
                cell.number_format = DATE_FORMAT


def format_sheet(worksheet, title: str, header_row: int = 3) -> None:
    """Format one report worksheet."""
    write_title(worksheet, title, worksheet.max_column)
    apply_table_formatting(worksheet, header_row)
    apply_number_formats(worksheet, header_row)
    apply_status_formatting(worksheet, header_row)


def build_final_report(output_file: Path = OUTPUT_FILE) -> None:
    """Create the final formatted Excel report workbook."""
    clean_data, sales_summary, quality_summary, issue_details = load_report_inputs()
    quality_summary = remove_duplicate_missing_checks(quality_summary)
    executive_summary = build_executive_summary(
        clean_data, quality_summary, issue_details
    )
    region_sales_summary = build_region_sales_summary(clean_data)

    output_file.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        friendly_columns(executive_summary).to_excel(
            writer, sheet_name="Executive Summary", index=False, startrow=2
        )
        friendly_columns(sales_summary["Revenue by Month"]).to_excel(
            writer, sheet_name="Monthly Sales", index=False, startrow=2
        )
        friendly_columns(sales_summary["Revenue by Category"]).to_excel(
            writer, sheet_name="Category Sales", index=False, startrow=2
        )
        friendly_columns(region_sales_summary).to_excel(
            writer, sheet_name="Region Sales", index=False, startrow=2
        )
        friendly_columns(sales_summary["Revenue by Sales Rep"]).to_excel(
            writer, sheet_name="Sales Rep Performance", index=False, startrow=2
        )
        friendly_columns(sales_summary["Top Products"]).to_excel(
            writer, sheet_name="Top Products", index=False, startrow=2
        )

        quality_sheet_name = "Data Quality"
        friendly_columns(quality_summary).to_excel(
            writer, sheet_name=quality_sheet_name, index=False, startrow=2
        )

        workbook = writer.book
        quality_sheet = workbook[quality_sheet_name]
        details_start_row = quality_sheet.max_row + 3
        quality_sheet.cell(row=details_start_row, column=1).value = "Issue Details"
        quality_sheet.cell(row=details_start_row, column=1).font = HEADER_FONT

        friendly_issue_details = friendly_columns(issue_details)
        for column_index, column_name in enumerate(friendly_issue_details.columns, start=1):
            cell = quality_sheet.cell(row=details_start_row + 1, column=column_index)
            cell.value = column_name
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        for row_offset, row_values in enumerate(
            friendly_issue_details.itertuples(index=False), start=details_start_row + 2
        ):
            for column_index, value in enumerate(row_values, start=1):
                quality_sheet.cell(row=row_offset, column=column_index).value = value

        for worksheet in workbook.worksheets:
            format_sheet(worksheet, worksheet.title)


def main() -> None:
    """Run the final report export."""
    build_final_report()
    print(f"Saved final sales report to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
